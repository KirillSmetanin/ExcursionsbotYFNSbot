import logging
import aiosqlite
from datetime import datetime, date
from telegram import Update, ReplyKeyboardMarkup, ReplyKeyboardRemove
from telegram.ext import (
    Application,
    CommandHandler,
    ConversationHandler,
    MessageHandler,
    filters,
    ContextTypes,
)
import re
import asyncio
import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from io import BytesIO

from config import BOT_TOKEN, WORKING_DAYS, WORKING_HOURS_START, WORKING_HOURS_END, DATE_FORMAT, TIME_FORMAT, DISPLAY_DATE_FORMAT, ERROR_MESSAGES
from database import db

# Включим логирование
logging.basicConfig(
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s", 
    level=logging.INFO,
    handlers=[
        logging.FileHandler('bot.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# Определим состояния диалога
(SCHOOL, CLASS, PROFILE, DATE, TIME, CONTACT_PERSON, 
 CONTACT_PHONE, PARTICIPANTS, CONFIRMATION) = range(9)

# Файл для хранения админов
ADMINS_FILE = 'admins.json'

# Загружаем список админов
def load_admins():
    """Загружаем список админов из файла"""
    try:
        if os.path.exists(ADMINS_FILE):
            with open(ADMINS_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
    except Exception as e:
        logger.error(f"Ошибка загрузки админов: {e}")
    return []

# Сохраняем список админов
def save_admins(admins_list):
    """Сохраняем список админов в файл"""
    try:
        with open(ADMINS_FILE, 'w', encoding='utf-8') as f:
            json.dump(admins_list, f, ensure_ascii=False, indent=2)
        return True
    except Exception as e:
        logger.error(f"Ошибка сохранения админов: {e}")
        return False

# Проверка является ли пользователь админом
def is_admin(user_id):
    """Проверяет, является ли пользователь админом"""
    admins = load_admins()
    return str(user_id) in admins

# Основное меню для админов
def get_main_menu_keyboard():
    """Основное меню для админов"""
    keyboard = [["📋 Забронировать экскурсию", "⚙️ Админ-панель"]]
    return ReplyKeyboardMarkup(keyboard, resize_keyboard=True)

# Клавиатура админ-панели
def get_admin_keyboard():
    """Клавиатура для админ-панели"""
    keyboard = [
        ["📊 Статистика", "📋 Все бронирования"],
        ["📅 Занятые даты", "📤 Экспорт в Excel"],
        ["👥 Управление админами", "📱 Отправить сообщение"],
        ["🔄 Очистить состояние", "🔙 В главное меню"]
    ]
    return ReplyKeyboardMarkup(keyboard, resize_keyboard=True)

# Клавиатура управления админами
def get_admin_management_keyboard():
    """Клавиатура для управления админами"""
    keyboard = [
        ["➕ Добавить админа", "➖ Удалить админа"],
        ["📋 Список админов", "🔙 Назад в админ-панель"]
    ]
    return ReplyKeyboardMarkup(keyboard, resize_keyboard=True)

# Функция-старт - упрощенная версия
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Начинаем диалог"""
    user = update.effective_user
    
    # Очищаем данные предыдущего диалога
    context.user_data.clear()
    
    # Проверяем админа
    if is_admin(user.id):
        # Админы видят меню выбора
        await update.message.reply_text(
            f"Здравствуйте, {user.first_name}! 👋\n"
            "Вы вошли как администратор.\n\n"
            "Выберите действие:",
            parse_mode='Markdown',
            reply_markup=get_main_menu_keyboard()
        )
        return ConversationHandler.END
    
    # Обычные пользователи сразу начинают бронирование
    await update.message.reply_text(
        f"Здравствуйте, {user.first_name}! 👋\n"
        "Этот бот поможет забронировать экскурсию для школьников в УФНС России по городу Москве.\n\n"
        "Пожалуйста, укажите полное название вашего учебного заведения, включая номер корпуса и фактический адрес:",
        parse_mode='Markdown',
        reply_markup=ReplyKeyboardRemove()
    )
    return SCHOOL

# Обработчик для названия школы
async def get_school(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Сохраняем название школы и спрашиваем класс"""
    school_name = update.message.text.strip()
    
    if len(school_name) < 3:
        await update.message.reply_text("Пожалуйста, введите полное название учебного заведения, включая номер корпуса и фактический адрес (минимум 3 символа):")
        return SCHOOL
    
    context.user_data['school'] = school_name
    await update.message.reply_text("Отлично! Теперь укажите класс (например, '10А' или '8'):")
    return CLASS

# Обработчик для класса
async def get_class(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Сохраняем класс и спрашиваем профильное направление"""
    class_number = update.message.text.strip()
    if not re.match(r'^[1-9][0-9]?[А-Яа-яA-Za-z]?$', class_number):
        await update.message.reply_text("Пожалуйста, введите корректный класс (например, '10А', '8Б' или '11'):")
        return CLASS
    
    context.user_data['class'] = class_number
    await update.message.reply_text(
        "Укажите профильное направление класса:\n"
        "Если профиля нет, напишите 'нет' или 'общеобразовательный'"
    )
    return PROFILE

# Обработчик для профиля
async def get_profile(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Сохраняем профиль и спрашиваем дату экскурсии"""
    profile = update.message.text.strip()
    context.user_data['profile'] = profile
    
    # Получаем забронированные даты для информации
    booked_dates = await db.get_booked_dates()
    booked_dates_str = ""
    if booked_dates:
        dates_formatted = []
        for d in booked_dates[:5]:
            try:
                date_obj = datetime.strptime(d, DATE_FORMAT)
                dates_formatted.append(date_obj.strftime(DISPLAY_DATE_FORMAT))
            except:
                continue
        booked_dates_str = "\n".join(dates_formatted)
    
    await update.message.reply_text(
        f"Профиль сохранен!\n\n"
        f"📅 *Теперь выберите дату экскурсии:*\n"
        f"• Введите дату в формате ДД.ММ.ГГГГ (например, 25.12.2024)\n"
        f"• Экскурсии проводятся только по вторникам, средам и четвергам!\n"
        f"• В один день может быть только одна экскурсия\n\n"
        f"📌 *Ближайшие занятые даты:*\n"
        f"{booked_dates_str if booked_dates_str else 'Нет занятых дат'}",
        parse_mode='Markdown',
        reply_markup=ReplyKeyboardRemove()
    )
    return DATE

# В обработчике get_date замените строку 210 на:
async def get_date(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Проверяем дату и спрашиваем время"""
    try:
        date_str = update.message.text.strip()
        try:
            excursion_date = datetime.strptime(date_str, "%d.%m.%Y").date()
        except ValueError:
            try:
                excursion_date = datetime.strptime(date_str, "%d/%m/%Y").date()
            except ValueError:
                excursion_date = datetime.strptime(date_str, "%Y-%m-%d").date()
        
        # Проверяем, что дата не в прошлом
        if excursion_date < date.today():
            await update.message.reply_text(ERROR_MESSAGES['date_passed'])
            return DATE
        
        # Проверяем день недели
        if excursion_date.weekday() not in WORKING_DAYS:
            await update.message.reply_text(ERROR_MESSAGES['invalid_day'])
            return DATE
        
        # Проверяем, занята ли дата
        try:
            is_date_available = await db.is_date_available(excursion_date.strftime(DATE_FORMAT))
        except Exception as e:
            logger.error(f"Ошибка проверки доступности даты: {e}")
            # Если функция не реализована, используем старый подход
            booked_times = await db.get_booked_slots_for_date(excursion_date.strftime(DATE_FORMAT))
            is_date_available = len(booked_times) == 0
        
        if not is_date_available:
            # Получаем информацию о существующей брони на эту дату
            try:
                booking_info = await db.get_booking_by_date(excursion_date.strftime(DATE_FORMAT))
                if booking_info:
                    # Форматируем информацию о занятой экскурсии
                    _, _, school, class_num, _, ex_date, ex_time, contact, _, participants, _ = booking_info
                    formatted_date = excursion_date.strftime(DISPLAY_DATE_FORMAT)
                    
                    await update.message.reply_text(
                        f"❌ *Дата {formatted_date} уже занята!*\n\n"
                        f"На эту дату уже запланирована экскурсия:\n"
                        f"• Школа: {school}\n"
                        f"• Класс: {class_num}\n"
                        f"• Время: {ex_time}\n"
                        f"• Контакт: {contact}\n"
                        f"• Участников: {participants}\n\n"
                        f"📌 *В один день может быть только одна экскурсия.*\n"
                        f"Пожалуйста, введите другую дату:",
                        parse_mode='Markdown'
                    )
                else:
                    await update.message.reply_text(
                        f"❌ Дата {excursion_date.strftime(DISPLAY_DATE_FORMAT)} уже занята.\n"
                        f"📌 В один день может быть только одна экскурсия.\n"
                        f"Пожалуйста, введите другую дату:"
                    )
            except Exception as e:
                logger.error(f"Ошибка получения информации о брони: {e}")
                await update.message.reply_text(
                    f"❌ Дата {excursion_date.strftime(DISPLAY_DATE_FORMAT)} уже занята.\n"
                    f"📌 В один день может быть только одна экскурсия.\n"
                    f"Пожалуйста, введите другую дату:"
                )
            return DATE
        
        # Сохраняем дату
        context.user_data['date'] = excursion_date.strftime(DATE_FORMAT)
        context.user_data['date_display'] = excursion_date.strftime(DISPLAY_DATE_FORMAT)
        
        await update.message.reply_text(
            f"✅ Дата {excursion_date.strftime(DISPLAY_DATE_FORMAT)} доступна!\n\n"
            f"⏰ *Введите время начала экскурсии:*\n"
            f"• Формат: ЧЧ:MM (например, 10:00)\n"
            f"• Время с {WORKING_HOURS_START}:00 до {WORKING_HOURS_END}:00",
            parse_mode='Markdown',
            reply_markup=ReplyKeyboardRemove()
        )
        return TIME
        
    except ValueError:
        await update.message.reply_text(
            "❌ Неверный формат даты!\n"
            "Пожалуйста, введите дату в формате ДД.ММ.ГГГГ (например, 25.12.2024):"
        )
        return DATE
    
# Обработчик для подтверждения (дополненная проверка)
async def confirmation(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Обрабатываем подтверждение или отмену заявки"""
    user_choice = update.message.text
    
    if user_choice == "✅ Подтвердить":
        user = update.effective_user
        
        try:
            required_fields = ['school', 'class', 'profile', 'date', 'time', 'contact_person', 'phone', 'participants']
            for field in required_fields:
                if field not in context.user_data:
                    await update.message.reply_text(
                        "❌ Не все данные заполнены. Пожалуйста, начните заново с /start",
                        reply_markup=ReplyKeyboardRemove()
                    )
                    context.user_data.clear()
                    return ConversationHandler.END
            
            # Двойная проверка доступности даты (на случай, если кто-то параллельно забронировал)
            is_date_available = await db.is_date_available(context.user_data['date'])
            if not is_date_available:
                # Получаем информацию о занятой дате
                booking_info = await db.get_booking_by_date(context.user_data['date'])
                if booking_info:
                    _, _, school, class_num, _, ex_date, ex_time, contact, _, participants, _ = booking_info
                    date_display = context.user_data['date_display']
                    
                    await update.message.reply_text(
                        f"❌ *Извините, эта дата только что занята!*\n\n"
                        f"Дата {date_display} теперь недоступна.\n"
                        f"На неё уже запланирована экскурсия:\n"
                        f"• Школа: {school}\n"
                        f"• Класс: {class_num}\n"
                        f"• Время: {ex_time}\n\n"
                        f"📌 *В один день может быть только одна экскурсия.*\n"
                        f"Пожалуйста, начните процесс заново с /start и выберите другую дату.",
                        parse_mode='Markdown',
                        reply_markup=ReplyKeyboardRemove()
                    )
                else:
                    await update.message.reply_text(
                        "❌ *Извините, эта дата только что занята!*\n\n"
                        "В один день может быть только одна экскурсия.\n"
                        "Пожалуйста, начните процесс заново с /start и выберите другую дату.",
                        parse_mode='Markdown',
                        reply_markup=ReplyKeyboardRemove()
                    )
                
                context.user_data.clear()
                return ConversationHandler.END
            
            # Сохраняем в базу данных
            success = await db.add_booking(
                user_id=user.id,
                username=user.username or f"{user.first_name} {user.last_name or ''}",
                school_name=context.user_data['school'],
                class_number=context.user_data['class'],
                class_profile=context.user_data['profile'],
                excursion_date=context.user_data['date'],
                excursion_time=context.user_data['time'],
                contact_person=context.user_data['contact_person'],
                contact_phone=context.user_data['phone'],
                participants_count=context.user_data['participants']
            )
            
            if success:
                await update.message.reply_text(
                    "🎉 *Поздравляем! Ваша заявка успешно оформлена!*\n\n"
                    f"📅 *Дата:* {context.user_data['date_display']}\n"
                    f"⏰ *Время:* {context.user_data['time']}\n\n"
                    "📞 С вами свяжется наш сотрудник для подтверждения деталей.\n"
                    "Чтобы создать новую заявку, нажмите /start",
                    parse_mode='Markdown',
                    reply_markup=ReplyKeyboardRemove()
                )
            else:
                await update.message.reply_text(
                    "❌ *Извините, произошла ошибка при сохранении!*\n\n"
                    "Пожалуйста, начните процесс заново с /start",
                    parse_mode='Markdown',
                    reply_markup=ReplyKeyboardRemove()
                )
                
        except Exception as e:
            logger.error(f"Ошибка сохранения заявки: {e}")
            await update.message.reply_text(
                "⚠️ Произошла ошибка при сохранении данных. Попробуйте позже.",
                reply_markup=ReplyKeyboardRemove()
            )
        
    else:  # Отмена
        await update.message.reply_text(
            "❌ Заявка отменена.\nЕсли хотите начать заново, используйте команду /start",
            reply_markup=ReplyKeyboardRemove()
        )
    
    context.user_data.clear()
    return ConversationHandler.END

# Обработчик для времени
async def get_time(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Проверяем время и спрашиваем контактное лицо"""
    time_str = update.message.text.strip()
    
    try:
        time_obj = datetime.strptime(time_str, TIME_FORMAT).time()
    except ValueError:
        await update.message.reply_text(
            "❌ Неверный формат времени!\n"
            f"Пожалуйста, введите время в формате ЧЧ:MM (например, 10:00):"
        )
        return TIME
    
    # Проверяем рабочее время
    if not (WORKING_HOURS_START <= time_obj.hour <= WORKING_HOURS_END):
        await update.message.reply_text(
            f"❌ Время должно быть с {WORKING_HOURS_START}:00 до {WORKING_HOURS_END}:00.\n"
            f"Пожалуйста, введите другое время:"
        )
        return TIME
    
    context.user_data['time'] = time_str
    
    await update.message.reply_text(
        "Отлично! Теперь укажите ФИО сопровождающего лица:",
        reply_markup=ReplyKeyboardRemove()
    )
    return CONTACT_PERSON

# Обработчик для контактного лица
async def get_contact_person(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Сохраняем контактное лицо и спрашиваем телефон"""
    contact_person = update.message.text.strip()
    if len(contact_person.split()) < 2:
        await update.message.reply_text("Пожалуйста, введите Фамилию и Имя (например, 'Иванов Иван'):")
        return CONTACT_PERSON
    
    context.user_data['contact_person'] = contact_person
    
    await update.message.reply_text(
        "Укажите контактный телефон для связи (в формате +7XXXXXXXXXX или 8XXXXXXXXXX):"
    )
    return CONTACT_PHONE

# Обработчик для телефона
async def get_contact_phone(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Проверяем телефон и спрашиваем количество участников"""
    phone = update.message.text.strip()
    
    # Очищаем телефон от лишних символов
    phone_clean = phone.replace(' ', '').replace('-', '').replace('(', '').replace(')', '')
    
    # Проверяем формат телефона
    phone_pattern = r'^(\+7|8|7)[\d]{10}$'
    if not re.match(phone_pattern, phone_clean):
        await update.message.reply_text(
            "❌ Неверный формат телефона!\n"
            "Пожалуйста, введите номер в формате +7XXXXXXXXXX или 8XXXXXXXXXX:"
        )
        return CONTACT_PHONE
    
    # Приводим к единому формату
    if phone_clean.startswith('8'):
        phone_clean = '+7' + phone_clean[1:]
    elif phone_clean.startswith('7'):
        phone_clean = '+' + phone_clean
    
    context.user_data['phone'] = phone_clean
    
    await update.message.reply_text(
        "Сколько всего участников планируется на экскурсии (школьники плюс не более 2 сопровождающих)?\n"
        "Введите число от 1 до 20:"
    )
    return PARTICIPANTS

# Обработчик для количества участников
async def get_participants(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Проверяем количество участников и показываем сводку"""
    try:
        participants = int(update.message.text.strip())
        
        if participants < 1 or participants > 20:
            await update.message.reply_text("Пожалуйста, введите число от 1 до 20:")
            return PARTICIPANTS
        
        context.user_data['participants'] = participants
        
        # Формируем сводку
        summary = (
            "📋 *Сводка вашей заявки:*\n\n"
            f"🏫 *Учебное заведение:* {context.user_data.get('school', 'Не указано')}\n"
            f"👨‍🎓 *Класс:* {context.user_data.get('class', 'Не указан')}\n"
            f"📚 *Профиль:* {context.user_data.get('profile', 'Не указан')}\n"
            f"📅 *Дата экскурсии:* {context.user_data.get('date_display', 'Не указана')}\n"
            f"⏰ *Время:* {context.user_data.get('time', 'Не указано')}\n"
            f"👤 *Сопровождающий:* {context.user_data.get('contact_person', 'Не указан')}\n"
            f"📞 *Телефон:* {context.user_data.get('phone', 'Не указан')}\n"
            f"👥 *Количество участников:* {context.user_data.get('participants', 'Не указано')}\n\n"
            "Всё верно?"
        )
        
        keyboard = [["✅ Подтвердить", "❌ Отмена"]]
        reply_markup = ReplyKeyboardMarkup(keyboard, one_time_keyboard=True, resize_keyboard=True)
        
        await update.message.reply_text(summary, parse_mode='Markdown', reply_markup=reply_markup)
        return CONFIRMATION
        
    except ValueError:
        await update.message.reply_text("Пожалуйста, введите число от 1 до 20:")
        return PARTICIPANTS

# Обработчик для подтверждения
async def confirmation(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Обрабатываем подтверждение или отмену заявки"""
    user_choice = update.message.text
    
    if user_choice == "✅ Подтвердить":
        user = update.effective_user
        
        try:
            required_fields = ['school', 'class', 'profile', 'date', 'time', 'contact_person', 'phone', 'participants']
            for field in required_fields:
                if field not in context.user_data:
                    await update.message.reply_text(
                        "❌ Не все данные заполнены. Пожалуйста, начните заново с /start",
                        reply_markup=ReplyKeyboardRemove()
                    )
                    context.user_data.clear()
                    return ConversationHandler.END
            
            # Двойная проверка доступности даты (на случай, если кто-то параллельно забронировал)
            is_date_available = await db.is_date_available(context.user_data['date'])
            if not is_date_available:
                await update.message.reply_text(
                    "❌ *Извините, эта дата только что занята!*\n\n"
                    "В один день может быть только одна экскурсия.\n"
                    "Пожалуйста, начните процесс заново с /start и выберите другую дату.",
                    parse_mode='Markdown',
                    reply_markup=ReplyKeyboardRemove()
                )
                context.user_data.clear()
                return ConversationHandler.END
            
            # Сохраняем в базу данных
            success = await db.add_booking(
                user_id=user.id,
                username=user.username or f"{user.first_name} {user.last_name or ''}",
                school_name=context.user_data['school'],
                class_number=context.user_data['class'],
                class_profile=context.user_data['profile'],
                excursion_date=context.user_data['date'],
                excursion_time=context.user_data['time'],
                contact_person=context.user_data['contact_person'],
                contact_phone=context.user_data['phone'],
                participants_count=context.user_data['participants']
            )
            
            if success:
                await update.message.reply_text(
                    "🎉 *Поздравляем! Ваша заявка успешно оформлена!*\n\n"
                    f"📅 *Дата:* {context.user_data['date_display']}\n"
                    f"⏰ *Время:* {context.user_data['time']}\n\n"
                    "📞 С вами свяжется наш сотрудник для подтверждения деталей.\n"
                    "Чтобы создать новую заявку, нажмите /start",
                    parse_mode='Markdown',
                    reply_markup=ReplyKeyboardRemove()
                )
            else:
                await update.message.reply_text(
                    "❌ *Извините, произошла ошибка при сохранении!*\n\n"
                    "Пожалуйста, начните процесс заново с /start",
                    parse_mode='Markdown',
                    reply_markup=ReplyKeyboardRemove()
                )
                
        except Exception as e:
            logger.error(f"Ошибка сохранения заявки: {e}")
            await update.message.reply_text(
                "⚠️ Произошла ошибка при сохранении данных. Попробуйте позже.",
                reply_markup=ReplyKeyboardRemove()
            )
        
    else:  # Отмена
        await update.message.reply_text(
            "❌ Заявка отменена.\nЕсли хотите начать заново, используйте команду /start",
            reply_markup=ReplyKeyboardRemove()
        )
    
    context.user_data.clear()
    return ConversationHandler.END

# Обработчик для команды отмены
async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Отменяет диалог"""
    await update.message.reply_text(
        "Диалог отменен. Если хотите начать заново, используйте команду /start",
        reply_markup=ReplyKeyboardRemove()
    )
    context.user_data.clear()
    return ConversationHandler.END

# Обработчик для команды help
async def help_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Показывает справку"""
    await update.message.reply_text(
        "📋 *Помощь по боту:*\n\n"
        "/start - Начать оформление заявки на экскурсию\n"
        "/help - Показать это сообщение\n"
        "/mybookings - Показать мои бронирования\n"
        "/cancel - Отменить текущий диалог\n\n"
        "*Важная информация:*\n"
        "• Экскурсии проводятся по вторникам, средам и четвергам\n"
        "• Время: с 10:00 до 15:00\n"
        "• В один день может быть только одна экскурсия\n"
        "• Максимальная группа: 20 человек (плюс не более 2 сопровождающих)",
        parse_mode='Markdown'
    )

# Обработчик для просмотра своих бронирований
async def my_bookings(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Показывает бронирования пользователя"""
    user = update.effective_user
    
    try:
        bookings = await db.get_user_bookings(user.id)
        
        if not bookings:
            await update.message.reply_text(
                "📭 У вас пока нет активных бронирований.\n"
                "Чтобы создать заявку, используйте команду /start"
            )
            return
        
        response = "📋 *Ваши активные бронирования:*\n\n"
        for i, booking in enumerate(bookings, 1):
            booking_id, school, class_num, ex_date, ex_time, contact, participants = booking
            try:
                date_formatted = datetime.strptime(ex_date, DATE_FORMAT).strftime(DISPLAY_DATE_FORMAT)
            except:
                date_formatted = ex_date
            
            response += (
                f"{i}. *ID:* {booking_id}\n"
                f"   🏫 {school}, класс {class_num}\n"
                f"   📅 {date_formatted} в {ex_time}\n"
                f"   👤 {contact}, 👥 {participants} чел.\n\n"
            )
        
        await update.message.reply_text(response, parse_mode='Markdown')
        
    except Exception as e:
        logger.error(f"Ошибка получения бронирований: {e}")
        await update.message.reply_text("⚠️ Произошла ошибка при получении данных.")

# ==================== АДМИН ФУНКЦИИ ====================

# Админ-панель
async def admin_panel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Показывает админ-панель"""
    user = update.effective_user
    
    if not is_admin(user.id):
        await update.message.reply_text("❌ У вас нет прав доступа.")
        return
    
    await update.message.reply_text(
        "⚙️ *Админ-панель*\n\nВыберите действие:",
        parse_mode='Markdown',
        reply_markup=get_admin_keyboard()
    )

# Показать статистику
async def admin_stats(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Показывает статистику"""
    user = update.effective_user
    
    if not is_admin(user.id):
        await update.message.reply_text("❌ У вас нет прав доступа.")
        return
    
    try:
        stats = await db.get_booking_stats()
        all_bookings = await db.get_all_bookings()
        
        days_stats = {0: 0, 1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0}
        for booking in all_bookings:
            try:
                ex_date = booking[5]
                date_obj = datetime.strptime(ex_date, DATE_FORMAT)
                day_of_week = date_obj.weekday()
                days_stats[day_of_week] += 1
            except:
                continue
        
        days_names = ["Пн", "Вт", "Ср", "Чт", "Пт", "Сб", "Вс"]
        days_stats_text = "\n".join([f"• {days_names[i]}: {days_stats[i]}" for i in WORKING_DAYS])
        
        response = (
            "📊 *Статистика бронирований*\n\n"
            f"📈 *Общая статистика:*\n"
            f"• Всего бронирований: {stats['total_bookings']}\n"
            f"• На будущее: {stats['future_bookings']}\n"
            f"• Всего участников: {stats['total_participants']}\n\n"
            f"📅 *По дням недели:*\n"
            f"{days_stats_text}"
        )
        
        await update.message.reply_text(response, parse_mode='Markdown')
        
    except Exception as e:
        logger.error(f"Ошибка получения статистики: {e}")
        await update.message.reply_text("❌ Ошибка при получении статистики.")

# Показать все бронирования
async def admin_all_bookings(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Показывает все бронирования"""
    user = update.effective_user
    
    if not is_admin(user.id):
        await update.message.reply_text("❌ У вас нет прав доступа.")
        return
    
    try:
        all_bookings = await db.get_all_bookings()
        
        if not all_bookings:
            await update.message.reply_text("📭 Нет активных бронирований.")
            return
        
        response = "📋 *Все бронирования:*\n\n"
        
        for booking in all_bookings:
            booking_id, username, school, class_num, profile, ex_date, ex_time, contact, phone, participants, booking_date = booking
            
            try:
                date_formatted = datetime.strptime(ex_date, DATE_FORMAT).strftime(DISPLAY_DATE_FORMAT)
            except:
                date_formatted = ex_date
            
            response += (
                f"🆔 *{booking_id}* | {date_formatted} {ex_time}\n"
                f"🏫 {school}, {class_num} ({profile})\n"
                f"👤 {contact} ({phone})\n"
                f"👥 {participants} чел. | 👤 {username if username else 'нет username'}\n\n"
            )
        
        # Разбиваем на части, если сообщение слишком длинное
        max_length = 4000
        if len(response) > max_length:
            parts = [response[i:i+max_length] for i in range(0, len(response), max_length)]
            for part in parts:
                await update.message.reply_text(part, parse_mode='Markdown')
        else:
            await update.message.reply_text(response, parse_mode='Markdown')
            
    except Exception as e:
        logger.error(f"Ошибка получения бронирований: {e}")
        await update.message.reply_text("❌ Ошибка при получении данных.")

# Показать занятые даты
async def admin_booked_dates(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Показывает занятые даты"""
    user = update.effective_user
    
    if not is_admin(user.id):
        await update.message.reply_text("❌ У вас нет прав доступа.")
        return
    
    try:
        booked_dates = await db.get_booked_dates()
        
        if not booked_dates:
            await update.message.reply_text("📅 Нет занятых дат.")
            return
        
        response = "📅 *Занятые даты:*\n\n"
        
        for date_str in booked_dates:
            try:
                date_obj = datetime.strptime(date_str, DATE_FORMAT)
                formatted_date = date_obj.strftime(DISPLAY_DATE_FORMAT)
                day_name = ["Пн", "Вт", "Ср", "Чт", "Пт", "Сб", "Вс"][date_obj.weekday()]
            except:
                formatted_date = date_str
                day_name = ""
            
            # Теперь на одну дату только одно время
            booking = await db.get_booking_by_date(date_str)
            if booking:
                response += f"• {formatted_date} ({day_name}): {booking[6]}\n"
            else:
                response += f"• {formatted_date} ({day_name})\n"
        
        await update.message.reply_text(response, parse_mode='Markdown')
        
    except Exception as e:
        logger.error(f"Ошибка получения занятых дат: {e}")
        await update.message.reply_text("❌ Ошибка.")

# Экспорт в Excel
async def admin_export_excel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Экспортирует данные в Excel"""
    user = update.effective_user
    
    if not is_admin(user.id):
        await update.message.reply_text("❌ У вас нет прав доступа.")
        return
    
    try:
        all_bookings = await db.get_all_bookings()
        
        if not all_bookings:
            await update.message.reply_text("📭 Нет данных для экспорта.")
            return
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Бронирования"
        
        headers = ["ID", "Дата брони", "ID пользователя", "Username", "Школа", "Класс", 
                   "Профиль", "Дата экскурсии", "Время", "Сопровождающий", "Телефон", "Количество"]
        
        # Записываем заголовки
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Записываем данные
        for row_idx, booking in enumerate(all_bookings, 2):
            (booking_id, username, school, class_num, profile, ex_date, ex_time, 
             contact_person, contact_phone, participants, booking_date) = booking
            
            ws.cell(row=row_idx, column=1, value=booking_id)
            ws.cell(row=row_idx, column=2, value=booking_date)
            ws.cell(row=row_idx, column=3, value=username)
            ws.cell(row=row_idx, column=4, value=username)
            ws.cell(row=row_idx, column=5, value=school)
            ws.cell(row=row_idx, column=6, value=class_num)
            ws.cell(row=row_idx, column=7, value=profile)
            ws.cell(row=row_idx, column=8, value=ex_date)
            ws.cell(row=row_idx, column=9, value=ex_time)
            ws.cell(row=row_idx, column=10, value=contact_person)
            ws.cell(row=row_idx, column=11, value=contact_phone)
            ws.cell(row=row_idx, column=12, value=participants)
        
        # Настраиваем ширину колонок
        column_widths = [8, 18, 12, 15, 25, 8, 20, 12, 8, 20, 15, 10]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + i)].width = width
        
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        filename = f"bookings_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        await update.message.reply_document(
            document=excel_buffer,
            filename=filename,
            caption=f"📊 Экспорт данных ({len(all_bookings)} записей)"
        )
        
        logger.info(f"Экспорт в Excel выполнен, {len(all_bookings)} записей")
        
    except Exception as e:
        logger.error(f"Ошибка экспорта в Excel: {e}")
        await update.message.reply_text("❌ Ошибка при экспорте данных в Excel.")

# Управление админами
async def admin_management(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Показывает меню управления админами"""
    user = update.effective_user
    
    if not is_admin(user.id):
        await update.message.reply_text("❌ У вас нет прав доступа.")
        return
    
    await update.message.reply_text(
        "👥 *Управление администраторами*\n\nВыберите действие:",
        parse_mode='Markdown',
        reply_markup=get_admin_management_keyboard()
    )

# Показать список админов
async def admin_list_admins(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Показывает список админов"""
    user = update.effective_user
    
    if not is_admin(user.id):
        await update.message.reply_text("❌ У вас нет прав доступа.")
        return
    
    admins = load_admins()
    
    if not admins:
        await update.message.reply_text("📭 Список администраторов пуст.")
        return
    
    response = "👥 *Список администраторов:*\n\n"
    
    for i, admin_id in enumerate(admins, 1):
        response += f"{i}. ID: {admin_id}\n"
    
    response += f"\nВсего администраторов: {len(admins)}"
    
    await update.message.reply_text(response, parse_mode='Markdown')

# Добавить админа
async def admin_add_admin(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Добавляет нового админа"""
    user = update.effective_user
    
    if not is_admin(user.id):
        await update.message.reply_text("❌ У вас нет прав доступа.")
        return
    
    try:
        new_admin_id = update.message.text.strip()
        
        if not new_admin_id.isdigit():
            await update.message.reply_text("❌ ID должен быть числом.")
            return
        
        admins = load_admins()
        
        if new_admin_id in admins:
            await update.message.reply_text(f"❌ Пользователь с ID {new_admin_id} уже является администратором.")
            return
        
        admins.append(new_admin_id)
        
        if save_admins(admins):
            await update.message.reply_text(f"✅ Пользователь с ID {new_admin_id} добавлен в список администраторов.")
            logger.info(f"Добавлен новый администратор: {new_admin_id}")
        else:
            await update.message.reply_text("❌ Ошибка при сохранении списка администраторов.")
        
    except Exception as e:
        logger.error(f"Ошибка добавления админа: {e}")
        await update.message.reply_text("❌ Ошибка при добавлении администратора.")

# Удалить админа
async def admin_remove_admin(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Удаляет админа"""
    user = update.effective_user
    
    if not is_admin(user.id):
        await update.message.reply_text("❌ У вас нет прав доступа.")
        return
    
    try:
        admin_to_remove = update.message.text.strip()
        
        if not admin_to_remove.isdigit():
            await update.message.reply_text("❌ ID должен быть числом.")
            return
        
        admins = load_admins()
        
        if admin_to_remove not in admins:
            await update.message.reply_text(f"❌ Пользователь с ID {admin_to_remove} не найден в списке администраторов.")
            return
        
        if admin_to_remove == str(user.id):
            await update.message.reply_text("❌ Вы не можете удалить себя из администраторов.")
            return
        
        admins.remove(admin_to_remove)
        
        if save_admins(admins):
            await update.message.reply_text(f"✅ Пользователь с ID {admin_to_remove} удален из списка администраторов.")
            logger.info(f"Удален администратор: {admin_to_remove}")
        else:
            await update.message.reply_text("❌ Ошибка при сохранении списка администраторов.")
        
    except Exception as e:
        logger.error(f"Ошибка удаления админа: {e}")
        await update.message.reply_text("❌ Ошибка при удалении администратора.")

# Отправить сообщение всем пользователям
async def admin_broadcast_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Начинает процесс рассылки сообщений"""
    user = update.effective_user
    
    if not is_admin(user.id):
        await update.message.reply_text("❌ У вас нет прав доступа.")
        return
    
    # Устанавливаем флаг, что ожидаем сообщение для рассылки
    context.user_data['awaiting_broadcast'] = True
    
    await update.message.reply_text(
        "📢 *Рассылка сообщения*\n\n"
        "Отправьте сообщение, которое хотите разослать всем пользователям:",
        parse_mode='Markdown',
        reply_markup=ReplyKeyboardRemove()
    )

async def start_booking_for_admin(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Запускает процесс бронирования для админов"""
    user = update.effective_user
    
    if not is_admin(user.id):
        return
    
    # Очищаем все данные
    context.user_data.clear()
    
    await update.message.reply_text(
        f"Здравствуйте, {user.first_name}! 👋\n"
        "Пожалуйста, укажите полное название учебного заведения, включая номер корпуса и фактический адрес::",
        reply_markup=ReplyKeyboardRemove()
    )
    
    # Устанавливаем состояние, что мы начинаем бронирование
    context.user_data['in_booking_process'] = True
    
    # Возвращаем состояние SCHOOL, чтобы запустить ConversationHandler
    return SCHOOL

# Обработчик текстовых сообщений для админов
async def handle_admin_text(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обрабатывает текстовые сообщения в админ-режиме"""
    user = update.effective_user
    
    if not is_admin(user.id):
        return
    
    text = update.message.text
    
    # Обработка основных команд админ-панели
    if text == "⚙️ Админ-панель":
        await admin_panel(update, context)
    
    elif text == "📊 Статистика":
        await admin_stats(update, context)
    
    elif text == "📋 Все бронирования":
        await admin_all_bookings(update, context)
    
    elif text == "📅 Занятые даты":
        await admin_booked_dates(update, context)
    
    elif text == "📤 Экспорт в Excel":
        await admin_export_excel(update, context)
    
    elif text == "👥 Управление админами":
        await admin_management(update, context)
    
    elif text == "📱 Отправить сообщение":
        await admin_broadcast_message(update, context)
    
    elif text == "➕ Добавить админа":
        await update.message.reply_text(
            "Отправьте ID пользователя, которого хотите сделать администратором:",
            reply_markup=ReplyKeyboardRemove()
        )
        context.user_data['awaiting_admin_id_add'] = True
    
    elif text == "➖ Удалить админа":
        await update.message.reply_text(
            "Отправьте ID администратора, которого хотите удалить:",
            reply_markup=ReplyKeyboardRemove()
        )
        context.user_data['awaiting_admin_id_remove'] = True
    
    elif text == "📋 Список админов":
        await admin_list_admins(update, context)
    
    elif text == "🔙 Назад в админ-панель":
        await admin_panel(update, context)
    
    elif text == "🔙 В главное меню":
        await update.message.reply_text(
            "Главное меню:",
            reply_markup=get_main_menu_keyboard()
        )
    elif text == "🔄 Очистить состояние":
        await clear_state_command(update, context)
    
    # Обработка специальных запросов
    elif context.user_data.get('awaiting_broadcast'):
        # Рассылка сообщения
        context.user_data.pop('awaiting_broadcast', None)
        
        try:
            async with aiosqlite.connect("excursions.db") as conn:
                cursor = await conn.execute("SELECT DISTINCT user_id FROM bookings")
                user_ids = await cursor.fetchall()
            
            if not user_ids:
                await update.message.reply_text("📭 Нет пользователей для рассылки.")
                return
            
            user_ids = [str(uid[0]) for uid in user_ids]
            success_count = 0
            
            await update.message.reply_text(f"📤 Отправка сообщения {len(user_ids)} пользователям...")
            
            for user_id in user_ids:
                try:
                    await context.bot.send_message(
                        chat_id=user_id,
                        text=f"📢 *Сообщение от администратора:*\n\n{text}",
                        parse_mode='Markdown'
                    )
                    success_count += 1
                    await asyncio.sleep(0.1)
                except Exception as e:
                    logger.error(f"Ошибка отправки пользователю {user_id}: {e}")
            
            await update.message.reply_text(
                f"✅ *Рассылка завершена*\n\n"
                f"• Успешно отправлено: {success_count}\n"
                f"• Всего пользователей: {len(user_ids)}",
                parse_mode='Markdown',
                reply_markup=get_admin_keyboard()
            )
            
        except Exception as e:
            logger.error(f"Ошибка рассылки: {e}")
            await update.message.reply_text("❌ Ошибка при рассылке сообщений.")
    
    elif context.user_data.get('awaiting_admin_id_add'):
        context.user_data.pop('awaiting_admin_id_add', None)
        await admin_add_admin(update, context)
        await admin_management(update, context)
    
    elif context.user_data.get('awaiting_admin_id_remove'):
        context.user_data.pop('awaiting_admin_id_remove', None)
        await admin_remove_admin(update, context)
        await admin_management(update, context)
    
    # Обработка начала бронирования для админа
    elif context.user_data.get('awaiting_school'):
        context.user_data.pop('awaiting_school', None)
        await get_school(update, context)

async def clear_state_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Очищает состояние пользователя - для тестирования"""
    user = update.effective_user
    
    # Только для админов
    if not is_admin(user.id):
        await update.message.reply_text("❌ Только для администраторов.")
        return
    
    # Очищаем ВСЕ данные пользователя
    context.user_data.clear()
    
    # Сбрасываем состояние чата
    chat_id = update.effective_chat.id
    if chat_id in context.chat_data:
        context.chat_data[chat_id] = {}
    
    # БЕЗ parse_mode='Markdown' или исправьте звездочки
    await update.message.reply_text(
        "✅ *Состояние полностью очищено!*\n\n"
        "Удалено:\n"
        "• Все временные данные (user_data)\n"
        "• Состояния диалога\n"
        "• Кеш клавиатур\n\n"
        "Теперь можно начать заново с /start",
        reply_markup=ReplyKeyboardRemove()
        # УБРАТЬ parse_mode='Markdown'
    )
    
    logger.info(f"Админ {user.id} очистил состояние")

# Команда для просмотра состояния
async def debug_state_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Показывает текущее состояние пользователя"""
    user = update.effective_user
    
    if not is_admin(user.id):
        await update.message.reply_text("❌ Только для администраторов.")
        return
    
    # Информация о состоянии
    user_data_info = "📊 *Текущее состояние (user_data):*\n"
    if context.user_data:
        for key, value in context.user_data.items():
            user_data_info += f"• {key}: {value}\n"
    else:
        user_data_info += "• Пусто\n"
    
    # Информация о пользователе
    user_info = (
        f"👤 *Информация:*\n"
        f"• ID: {user.id}\n"
        f"• Админ: {'✅ Да' if is_admin(user.id) else '❌ Нет'}\n"
    )
    
    response = user_info + "\n" + user_data_info
    
    await update.message.reply_text(response, parse_mode='Markdown')

# Обработчик ошибок
async def error_handler(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Логирует ошибки"""
    logger.error(f"Ошибка: {context.error}", exc_info=context.error)
    
    if update and update.effective_message:
        await update.effective_message.reply_text(
            "⚠️ Произошла непредвиденная ошибка. Попробуйте позже или начните заново с /start"
        )

async def main() -> None:
    """Асинхронный запуск бота"""
    # Инициализируем базу данных
    await db.init_db()
    logger.info("База данных инициализирована")
    
    # Создаем Application
    application = Application.builder().token(BOT_TOKEN).build()

    # Создаем ConversationHandler для основного диалога (бронирования)
    conv_handler = ConversationHandler(
        entry_points=[
            CommandHandler("start", start),
            MessageHandler(filters.Regex(r'^📋 Забронировать экскурсию$') & filters.TEXT, start_booking_for_admin)
        ],
        states={
            SCHOOL: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_school)],
            CLASS: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_class)],
            PROFILE: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_profile)],
            DATE: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_date)],
            TIME: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_time)],
            CONTACT_PERSON: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_contact_person)],
            CONTACT_PHONE: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_contact_phone)],
            PARTICIPANTS: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_participants)],
            CONFIRMATION: [MessageHandler(filters.TEXT & ~filters.COMMAND, confirmation)],
        },
        fallbacks=[
            CommandHandler("cancel", cancel),
            CommandHandler("help", help_command),
        ],
        allow_reentry=True,
    )

    # Добавляем обработчики
    application.add_handler(conv_handler)
    application.add_handler(CommandHandler("help", help_command))
    application.add_handler(CommandHandler("mybookings", my_bookings))
    application.add_handler(CommandHandler("admin", admin_panel))
    application.add_handler(CommandHandler("cancel", cancel))

        # Добавляем обработчики команд (ДОБАВЬТЕ ЭТИ ДВЕ СТРОЧКИ):
    application.add_handler(CommandHandler("clear", clear_state_command))  # Очистка состояния
    application.add_handler(CommandHandler("debug", debug_state_command))  # Просмотр состояния
    
    # Обработчик для текстовых сообщений админов
    application.add_handler(MessageHandler(
        filters.TEXT & ~filters.COMMAND,
        handle_admin_text
    ))
    
    # Обработчик ошибок
    application.add_error_handler(error_handler)

    # Создаем файл админов при первом запуске, если его нет
    if not os.path.exists(ADMINS_FILE):
        initial_admin_id = "ВАШ_TELEGRAM_ID"  # ЗАМЕНИТЕ НА ВАШ ID
        save_admins([initial_admin_id])
        logger.info(f"Создан файл админов, добавлен администратор с ID: {initial_admin_id}")
    
    # Запускаем бота
    logger.info("Бот запускается...")
    await application.initialize()
    await application.start()
    await application.updater.start_polling()
    
    # Ждем сигнала остановки
    try:
        while True:
            await asyncio.sleep(1)
    except KeyboardInterrupt:
        logger.info("Остановка бота...")
    finally:
        await application.stop()
        logger.info("Бот остановлен")

if __name__ == "__main__":
    asyncio.run(main())